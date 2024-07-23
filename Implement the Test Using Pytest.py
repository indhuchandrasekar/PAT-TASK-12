import pytest
from selenium import webdriver
import openpyxl
from datetime import datetime

@pytest.fixture(scope="module")
def setup():
    driver = webdriver.Chrome()
    driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")
    yield driver
    driver.quit()

def read_test_data(file_name):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active
    test_data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        test_data.append(row)
    return test_data, workbook, sheet

def test_login(setup):
    driver = setup
    login_page = LoginPage(driver)
    test_data, workbook, sheet = read_test_data("test_data.xlsx")

    for index, data in enumerate(test_data, start=2):
        test_id, username, password, date, time, tester, result = data
        login_page.login(username, password)
        try:
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, "some_element_id_after_login"))
            )
            result = "Passed"
        except:
            result = "Failed"

        sheet.cell(row=index, column=7).value = result
        workbook.save("test_data.xlsx")
